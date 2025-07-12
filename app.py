from flask import Flask, render_template, request, url_for
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

# ✅ Read brand name from business.txt
def get_business_name():
    try:
        with open('business.txt', 'r', encoding='utf-8') as f:
            return f.readline().strip()
    except FileNotFoundError:
        return "My Company"

# ✅ Read about us text from about.txt
def get_about_info():
    try:
        with open('about.txt', 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        return ""

# ✅ Read models for a product from models/{product}.txt
def load_models_for_product(product):
    models = []
    model_file = os.path.join("models", f"{product}.txt")
    if os.path.exists(model_file):
        with open(model_file, 'r', encoding='utf-8') as f:
            for line in f:
                model = line.strip()
                if model:
                    # ✅ Updated folder name to model_img
                    png_path = f"static/model_img/{model}.png"
                    jpg_path = f"static/model_img/{model}.jpg"
                    if os.path.exists(png_path):
                        image = f"model_img/{model}.png"
                    elif os.path.exists(jpg_path):
                        image = f"model_img/{model}.jpg"
                    else:
                        image = None
                    models.append({'name': model, 'image': image})
    return models

@app.route('/')
def form():
    product_file = 'products.txt'
    products = []

    if os.path.exists(product_file):
        with open(product_file, 'r', encoding='utf-8') as f:
            for line in f:
                product = line.strip()
                if product:
                    models = load_models_for_product(product)
                    products.append({'name': product, 'models': models})

    business_name = get_business_name()
    logo_file = "logo.png"
    about_info = get_about_info()

    return render_template(
        "form.html",
        products=products,
        business_name=business_name,
        logo_file=logo_file,
        about_info=about_info
    )

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name'].strip()
    business = request.form['business'].strip() or "NA"
    phone = request.form['phone'].strip()
    email = request.form.get('email', '')
    month = request.form.get('month', '')
    selected_models = request.form.getlist('models')

    # Validation
    error = None
    if not name.replace(" ", "").isalnum():
        error = "Name must contain only alphanumeric characters and spaces."
    elif not phone.isdigit() or len(phone) != 10:
        error = "Phone number must be exactly 10 digits."

    if error:
        # Reload products so we can re-render form
        product_file = 'products.txt'
        products = []
        if os.path.exists(product_file):
            with open(product_file, 'r', encoding='utf-8') as f:
                for line in f:
                    product = line.strip()
                    if product:
                        models = load_models_for_product(product)
                        products.append({'name': product, 'models': models})

        business_name = get_business_name()
        logo_file = "logo.png"
        about_info = get_about_info()

        return render_template(
            "form.html",
            error=error,
            products=products,
            business_name=business_name,
            logo_file=logo_file,
            about_info=about_info,
            prefill={
                'name': name,
                'business': business,
                'phone': phone,
                'email': email,
                'month': month,
                'selected_models': selected_models
            }
        )

    # Set default for business if empty
    business = business or "NA"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    selected_models_str = ', '.join(selected_models)

    # Save to Excel
    filename = 'submissions.xlsx'
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Enquiries"
        ws.append(['Name', 'Business Name', 'Phone', 'Email', 'Product(s) + Models', 'Preferred Month', 'Timestamp'])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    ws.append([name, business, phone, email, selected_models_str, month, timestamp])
    wb.save(filename)

    return "Thank you! Your enquiry has been recorded."

if __name__ == '__main__':
    app.run(debug=True)
